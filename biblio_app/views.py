from django.shortcuts import render, redirect
from django.contrib import admin, messages
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import authenticate, login
from .models import *
from django.db.models import Q, CharField, Value, F
import pandas as pd
from datetime import datetime, timedelta
from django.core.exceptions import ValidationError
from django.utils import timezone
from openpyxl import Workbook
from docx import Document
from docx.shared import Inches
from docxtpl import DocxTemplate
from django.db.models.functions import Concat
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.contrib.auth.decorators import login_required

listacopiasfront=[]
listacopiastras=[]
listacopiasint=[]
listacredenciales=[]
vaciocopias=[]
vacioalum=[]
alumnoult=[]
libroult=[]
copiault=[]
prestault=[]

class auxiliar:
    def __init__(self, modelo_existente, campo_auxiliar):
        self.copia = modelo_existente
        self.campo_auxiliar = campo_auxiliar


def get_server_time(request):
    # Obtiene la hora actual del servidor
    server_time = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Devuelve la hora del servidor como una respuesta JSON
    return JsonResponse({'server_time': server_time})

#register your models here
@login_required
def alumnoagre_pest(request):
    alumnos = Alumno.objects.all()
    # Devolver una respuesta con la lista de alumnos encontrados
    ultimo=alumnoult[0]
    return render(request, 'alumnos.html', {"current_tab": "alumno", "alumnos": alumnos,"ultimo": ultimo})

@login_required
def alumnoagre_pest_excel(request):
    ultimo=alumnoult
    return render(request, 'agregado_excel_alumnos.html', {"alumnos": ultimo})

@login_required
def copiaagregada(request):
    libros = Libro.objects.all()
    ultimacopia=copiault[0]
    return render(request, "libros.html", context={"current_tab": "libro", "libros": libros,"copiault":ultimacopia})

@login_required
def copiaagregadaexcel(request):
    ultimacopia=copiault
    return render(request, "agregado_excel_copias.html", context={"copias":ultimacopia})
    

@login_required
def libroagregado(request):
    libros = Libro.objects.all()
    ultimolibro=libroult[0]
    return render(request, "libros.html", context={"current_tab": "libro", "libros": libros,"libroult":ultimolibro})

@login_required
def libroagregadoexcel(request):
    ultimolibro=libroult
    return render(request, "agregado_excel_libros.html", context={"libros":ultimolibro})


def nuevoprestamo(request):
    prestamo = Prestamo.objects.all()
    if prestault:
        ultimo=prestault[0]
    else:
        ultimo=None
    return render(request, "prestamo.html", context={"current_tab": "prestamo", "prestamo": prestamo,"ultimo":ultimo})


# Create your views here.
def inicio(request):
    return render(request,"inicio.html",context={"current_tab": "inicio"})


def prestamos(request):
    query = request.GET.get('q')
    if query:
        # Verificar si la consulta es un número puro
        if query.isdigit():
            # Eliminar los ceros a la izquierda en el caso de que existan
            query = str(int(query))
        
        # Realizar la búsqueda utilizando el campo 'clave alumno o clave copia'
        prestamo = Prestamo.objects.filter(
            Q(clave_alumno__clave__icontains=query) |
            Q(clave_copia__clave__icontains=query)
        ).distinct()
    else:
        # Si no hay búsqueda, mostrar todos los prestamos
        prestamo = Prestamo.objects.all()

    # Ordenar los préstamos del más nuevo al más antiguo
    prestamo = prestamo.order_by(F('fecha_creacion').desc(nulls_last=True))

    # Devolver una respuesta con la lista de préstamos ordenados
    return render(request, "prestamo.html", context={"current_tab": "prestamo", "prestamo": prestamo})


def retornos(request):
    query = request.GET.get('q')
    if query:
        # Verificar si la consulta es un número puro
        if query.isdigit():
            # Eliminar los ceros a la izquierda en el caso de que existan
            query = str(int(query))
        
        # Realizar la búsqueda utilizando el campo 'clave alumno o clave copia'
        prestamo = Prestamo.objects.filter(
            Q(clave_alumno__clave__icontains=query) |
            Q(clave_copia__clave__icontains=query) &
            Q(activo=True)
        ).distinct()
    else:
        # Si no hay búsqueda, mostrar todos los préstamos activos
        prestamo = Prestamo.objects.filter(activo=True)

    # Devolver una respuesta con la lista de préstamos
    return render(request,"retorno.html",context={"current_tab": "retorno", "prestamo": prestamo})

def multas(request):
    query = request.GET.get('q')
    if query:
        # Verificar si la consulta es un número puro
        if query.isdigit():
            # Eliminar los ceros a la izquierda en el caso de que existan
            query = str(int(query))
        
        # Realizar la búsqueda utilizando el campo 'clave alumno o clave copia'
        multas = Multa.objects.filter(
            Q(clave_alumno_clave__icontains=query) &
            Q(pagado=False) &
            Q(monto__gt=10)  # Filtrar multas con monto mayor a 10
        ).distinct()
    else:
        # Si no hay búsqueda, mostrar todos los préstamos activos
        multas = Multa.objects.filter(pagado=False, monto__gt=9)  # Filtrar multas con monto mayor a 10

    # Devolver una respuesta con la lista de préstamos
    return render(request, "multas.html", context={"current_tab": "multa", "multas": multas})


def etiquetas(request):
    listacopias=[]
    for i in vaciocopias:
        listacopias.append(i)
    
    for i in listacopiasfront:
        listacopias.append(i)
    for i in listacopiasint:
        listacopias.append(i)
    for i in listacopiastras:
        listacopias.append(i)

    listacopias=ordena(listacopias)
    return render(request,"etiqueta.html",context={"current_tab": "etiqueta", "lista": listacopias, "vacios":vaciocopias})
def credenciales(request):
    return render(request,"credencial.html",context={"current_tab": "credencial", "lista": listacredenciales,"vacios":vacioalum})

def ordena(x):
    y=[]
    z=[]
    for i in x:
        if i.copia == '':
            y.append(i)
        else:
            z.append(i)
    z.sort(key=lambda x: x.copia.clavecopia)
    lista=[]
    for i in y:
        lista.append(i)
    for i in z:
        lista.append(i)
    return(lista)

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('')  # Redirige a la URL deseada después del inicio de sesión
        else:
            # Lógica para manejar credenciales incorrectas
            messages.error(request, 'Credenciales incorrectas. Por favor, inténtalo de nuevo.')
            pass

    return render(request, 'login.html')

def alumno_pest(request):
    query = request.GET.get('q')
    if query:
        # Verificar si la consulta es un número puro
        if query.isdigit():
            # Eliminar los ceros a la izquierda en el caso de que existan
            query = str(int(query))
        
        # Realizar la búsqueda utilizando el campo 'nombre', 'apellido', 'clave' o 'grupo'
        alumnos = Alumno.objects.annotate(
            nombre_apellido=Concat('nombre', Value(' '), 'apellido', output_field=CharField()),
            grado_grupo=Concat('grupo','clase',output_field=CharField()),
        ).filter(
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query) |
            Q(clave__icontains=query) |
            Q(grupo__icontains=query) |
            Q(grado_grupo__icontains=query) |
            Q(nombre_apellido__icontains=query)  # Buscar en el campo combinado de nombre y apellido
        ).distinct()
        
    else:
        # Si no hay búsqueda, mostrar todos los alumnos
        alumnos = Alumno.objects.all()

    # Devolver una respuesta con la lista de alumnos encontrados
    return render(request, 'alumnos.html', {"current_tab": "alumno", "alumnos": alumnos})

@login_required
def agregar_alum(request):
    if request.method == 'POST':
        # Obtener los datos del formulario
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        grupo = request.POST.get('grupo')
        clase= request.POST.get('clase','')

        # Verificar si los campos obligatorios no están vacíos
        if not nombre or not apellido or not grupo:
            return render(request, 'error.html', {'mensaje': 'Por favor, completa todos los campos obligatorios.'})

        try:
            # Intentar crear una instancia de Alumno con los datos proporcionados
            alumno_item = Alumno(nombre=nombre, apellido=apellido, grupo=grupo,clase=clase)
            alumno_item.full_clean()  # Validar los datos del modelo
            alumno_item.save()  # Guardar el objeto en la base de datosalumnos = Alumno.objects.all() #carga base de alumnos
            alumnoult.clear()
            alumnoult.append(alumno_item)
            return redirect('/alumno_agregado')
        
        except ValidationError as e:
            # Si se produce una excepción de validación, mostrar un mensaje de error
            return render(request, 'error.html', {'mensaje': '; '.join(e.messages)})

    return render(request, 'formulario_alumno.html')

def alumno_detalle(request, pk):
    alumno_obj = get_object_or_404(Alumno, clave=pk)
    return render(request, 'alumno_detalle.html', {'alumno': alumno_obj})

@login_required #verifica que este iniciada sesion
def editar_alumno(request, pk):
    alumno_obj = get_object_or_404(Alumno, clave=pk)

    if request.method == 'POST':
        alumno_obj.nombre = request.POST['nombre']
        alumno_obj.apellido = request.POST['apellido']
        alumno_obj.grupo = request.POST['grupo']
        alumno_obj.clase = request.POST['clase']
        alumno_obj.save()
        return redirect('/alumno')

    return render(request, 'editar_alumno.html', {'alumno': alumno_obj})

@login_required #verifica que este iniciada sesion
def eliminar_alumno(request, pk):
    alumno_obj = get_object_or_404(Alumno, clave=pk)
    print(alumno_obj)

    if request.method == 'POST':
        alumno_obj.delete()
        return redirect('/alumno')

    return render(request, 'borrar_alumno.html', {'alumno': alumno_obj})

@login_required #verifica que este iniciada sesion
def cargar_desde_excel(request):
    if request.method == 'POST' and 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        df = pd.read_excel(excel_file)

        required_columns = ['nombre', 'apellido', 'grado']
        missing_columns = [col for col in required_columns if col not in df.columns]
       

        if missing_columns:
            missing_columns_str = ', '.join(missing_columns)
            error_message = f"El archivo Excel no tiene las columnas necesarias: {missing_columns_str}."
            return render(request, 'error.html', {'mensaje': error_message})
        alumnoult.clear()
        for index, row in df.iterrows():
            if 'clave' in df.columns and not pd.isna(row['clave']):
                clave = row['clave']
            else:
                clave = None  # Asignar None para que se autoasigne la clave por el sistema

            alumno_item = Alumno(
                clave=clave,
                nombre=row['nombre'],
                apellido=row['apellido'],
                grupo=row['grado'],
                clase=row.get('grupo', None),
            )

            try:
                alumno_item.full_clean()  # Validar los datos del modelo
                alumno_item.save()  # Guardar el objeto en la base de datos
                alumnoult.append(alumno_item)
            except ValidationError as e:
                error_message = '; '.join(e.messages)

        return redirect('/alumno_agregado_excel')

    return render(request, 'tu_template_excel.html')


@login_required #verifica que este iniciada sesion
def borrar_todos_los_alumnos(request):
    try:
        # Verificar si existen alumnos con grupo igual a 6 y sacalibro verdadero
        if Alumno.objects.filter(grupo=6, sacalibro=True).exists():
            # Si existen, lanzar un error
              return render(request, 'error.html', {'mensaje': 'Existen alumnos de 6to con préstamos activos. Favor de completar estos préstamos para realizar el proceso de final de año.'})
       
        # Eliminar todos los alumnos con un valor de grupo de 6 o superior
        Alumno.objects.filter(grupo__gte=6).delete()
        
        # Modificar todos los datos en la tabla de alumnos sumando uno al valor de grupo
        Alumno.objects.all().update(grupo=models.F('grupo') + 1)
        
        # Redireccionar a la página de alumnos
        return redirect('/alumno')  # Cambia 'alumnos' por la URL de la página de alumnos en tu proyecto
    
    except ValidationError as e:
        # Si se produce una excepción de validación, mostrar un mensaje de error
        return render(request, 'error.html', {'mensaje': '; '.join(e.messages)})
    
def libros_pest(request):
    query = request.GET.get('q')
    if query:
        # Verificar si la consulta es un número puro
        if query.isdigit():
            # Eliminar los ceros a la izquierda en el caso de que existan
            query = str(int(query))

        libros = Libro.objects.filter(
            Q(codigolibro__icontains=query) |
            Q(titulo__icontains=query) |
            Q(autor__icontains=query) |
            Q(editorial__icontains=query)|
            Q(palabrasclave__icontains=query)
        )
    else:
        libros = Libro.objects.all()

    return render(request, "libros.html", context={"current_tab": "libro", "libros": libros})

@login_required #verifica que este iniciada sesion
def agregar_libros(request):
    if request.method == 'POST':
        try:
            fechapubl_str = request.POST.get('fechapubl', '')
            if fechapubl_str:
                fechapubl = int(fechapubl_str)
            else:
                fechapubl = None

            libro_item = Libro(
                titulo=request.POST.get('titulo', ''),
                autor=request.POST.get('autor', ''),
                editorial=request.POST.get('editorial', ''),
                numerotomo=request.POST.get('numtomo', ''),
                dewy=request.POST.get('ubicacion', ''),  # Ahora 'dewy' en lugar de 'ubicacion'
                publicodirigido=request.POST.get('publico', ''),
                fechapublicacion=fechapubl,
                caracteristicasespeciales=request.POST.get('caracteristicas', ''),
                ilustrador=request.POST.get('ilustrador',''),
                palabrasclave=request.POST.get('clave',''),
                item=request.POST.get('item','')
                )
            libro_item.full_clean()  # Realiza todas las validaciones del modelo
            libro_item.save()
            libroult.clear()
            libroult.append(libro_item)
            return redirect('/libro_agregado')
        
        except ValidationError as e:
            # Si se produce una excepción de validación, mostrar un mensaje de error
            return render(request, 'error.html', {'mensaje': '; '.join(e.messages)})

    return redirect('/libro')

@login_required #verifica que este iniciada sesion
def agregar_copia(request):
    if request.method == 'POST':
        # Obtener el código del libro del formulario
        codigolibro = request.POST.get('codigolibro')
        
        # Verificar si el libro con ese código existe en la base de datos
        libro = get_object_or_404(Libro, codigolibro=codigolibro)
        
        # Si el libro existe, crear una instancia de Copia utilizando el libro obtenido
        if libro:
            copia_item = Copia(codigolibro=libro)
            
            # Guardar la instancia de Copia en la base de datos
            copia_item.save()
            copiault.clear()
            copiault.append(copia_item)
            # Redireccionar a la página de libros
            return redirect('/copia_agregada')
        else:
            # Si el libro no existe, puedes mostrar un mensaje de error o redireccionar a una página de error
           return render(request, 'error.html', {'mensaje': 'El codigo de libro introducido no corresoponde an ningun libro en la base favor de verificarlo'})
    else:
        # Si la solicitud no es POST, probablemente deberías manejarla de alguna manera
        # Por ejemplo, podrías mostrar un mensaje de error o redirigir a una página de error
        return redirect('/')


def libro_detalle(request, codigolibro):
    libro_obj = get_object_or_404(Libro, codigolibro=codigolibro)
    return render(request, 'libros_detalle.html', {'libro': libro_obj})

@login_required #verifica que este iniciada sesion
def editar_libro(request, codigolibro):
    libro_obj = get_object_or_404(Libro, codigolibro=codigolibro)

    if request.method == 'POST':
        # Procesa el formulario de edición aquí
        libro_obj.titulo = request.POST['titulo']
        libro_obj.autor = request.POST['autor']
        libro_obj.editorial = request.POST['editorial']

        # Los campos no obligatorios se manejan si están presentes en la solicitud POST
        libro_obj.ilustrador = request.POST.get('ilustrador', None)
        libro_obj.numerotomo = request.POST.get('numerotomo', None)
        libro_obj.caracteristicasespeciales = request.POST.get('caracteristicasespeciales', None)
        libro_obj.dewy = request.POST.get('ubicacionbiblio', None)
        libro_obj.publicodirigido = request.POST.get('publicodirigido', None)
        libro_obj.palabrasclave = request.POST.get('clave', None)
        libro_obj.item = request.POST.get('item', None)

        # Obtener el valor de fechapublicacion del formulario
        fechapublicacion = request.POST.get('fechapublicacion')

        # Si el valor obtenido es None o vacío, establecer fechapublicacion como None
        if not fechapublicacion:
            libro_obj.fechapublicacion = None
        else:
            try:
                # Intentar convertir el valor a un entero
                fechapublicacion = int(fechapublicacion)
                libro_obj.fechapublicacion = fechapublicacion
            except ValueError:
                error_message = 'Formato de fecha de publicación no válido. Por favor, ingresa un entero válido.'
                return render(request, 'error.html', {'mensaje': error_message})

        libro_obj.save()
        return redirect('/libro')

    # Si la solicitud no es POST, renderiza el formulario de edición
    return render(request, 'editar_libro.html', {'libro': libro_obj})


@login_required #verifica que este iniciada sesion
def eliminar_libro(request, codigolibro):
    libro_obj = get_object_or_404(Libro, codigolibro=codigolibro)
    if request.method == 'POST':
        libro_obj.delete()
        return redirect('/libro')
    return render(request, 'eliminar_libro.html', {'libro': libro_obj})

@login_required #verifica que este iniciada sesion
def cargar_desde_excel_libro(request):
    if request.method == 'POST':
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)

            required_columns = ['titulo', 'autor', 'editorial']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                missing_columns_str = ', '.join(missing_columns)
                error_message = f"El archivo Excel no tiene las columnas necesarias: {missing_columns_str}."
                return render(request, 'error.html', {'mensaje': error_message})
            
            libroult.clear()
            for index, row in df.iterrows():
                if any(pd.isna(row[col]) for col in required_columns):
                    error_message = "Los campos obligatorios no pueden estar vacíos."
                    return render(request, 'error.html', {'mensaje': error_message})

                if 'codigolibro' in df.columns and pd.notna(row['codigolibro']):
                    codigolibro = row['codigolibro']
                else:
                    codigolibro = None

                fechapublicacion = int(row.get('fechapublicacion', 0))  # Cambiar a entero

                libro_item = Libro(
                    codigolibro=codigolibro,
                    titulo=row['titulo'],
                    autor=row['autor'],
                    editorial=row['editorial'],
                    ilustrador=row.get('ilustrador', ''),
                    fechapublicacion=fechapublicacion,
                    numerotomo=row.get('tomo', ''),
                    caracteristicasespeciales=row.get('caracteristicas', ''),
                    dewy=row.get('ubicacionbiblio', ''),
                    publicodirigido=row.get('publico', ''),
                    item=row.get('item', ''),
                    palabrasclave=row.get('palabras_clave', ''),
                )
                print(libro_item)
                try:
                    libro_item.full_clean()
                    libro_item.save()
                    libroult.append(libro_item)
                except ValidationError as e:
                    error_message = '; '.join(e.messages)
                    print(error_message)
                    

            return redirect('/libro_agregado_excel')  # Redirigir a la página de libros después de cargar las copias

    return render(request, 'tu_template_excel.html')

@login_required #verifica que este iniciada sesion
def borrar_todos_los_libros(request):
    try:
        # Eliminar todos los libros
        Libro.objects.all().delete()
        return redirect('/libro')
    except Exception as e:
        error_message = 'Se produjo un error al intentar borrar todos los libros. Por favor, inténtalo de nuevo.'
        return render(request, 'error.html', {'mensaje': error_message})

@login_required #verifica que este iniciada sesion
def eliminar_libros_por_titulo(request):
    if request.method == 'POST':
        titulo = request.POST.get('titulo', None)
        if titulo:
            Libro.objects.filter(titulo=titulo).delete()
    return redirect('/libro')
    

def busqueda_pest(request):
    query = request.GET.get('q')
    if query:
        # Verificar si la consulta es un número puro
        if query.isdigit():
            # Eliminar los ceros a la izquierda en el caso de que existan
            query = str(int(query))
    if query:
        copias = Copia.objects.filter(Q(clavecopia__icontains=query) | Q(codigolibro__titulo__icontains=query) | Q(codigolibro__autor__icontains=query) | Q(codigolibro__editorial__icontains=query) | Q(codigolibro__ilustrador__icontains=query)|Q(codigolibro__palabrasclave__icontains=query))
    else:
        # Si no hay búsqueda, muestra todas las copias
        copias = Copia.objects.all()

    return render(request,"busqueda.html",context={"current_tab": "busqueda", "copias": copias})


def busqeda_detalle(request, clavecopia):
    # Usamos get_object_or_404 para obtener el objeto alumno o retornar un error 404 si no se encuentra
    copia_obj = get_object_or_404(Copia, clavecopia=clavecopia)
    
    return render(request, 'busquedadetalle.html', {'copia': copia_obj})

@login_required #verifica que este iniciada sesion
def cargar_copias_desde_excel(request):
    if request.method == 'POST':
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)

            required_columns = ['codigolibro']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                missing_columns_str = ', '.join(missing_columns)
                error_message = f"El archivo Excel no tiene las columnas necesarias: {missing_columns_str}."
                return render(request, 'error.html', {'mensaje': error_message})
            copiault.clear()
            try:
                for index, row in df.iterrows():
                    codigolibro = row['codigolibro']

                    if 'clavecopia' in df.columns and not pd.isna(row['clavecopia']):
                        clavecopia = row['clavecopia']
                    else:
                        clavecopia = None  # Asignar None para que se autoasigne la clave de copia por Django

                    libro_existente = Libro.objects.filter(codigolibro=codigolibro).first()
                    if libro_existente:
                        copia_item = Copia(
                            clavecopia=clavecopia,
                            codigolibro=libro_existente,
                            disponible=True,  # Asignar True por defecto
                            clavealumno=None,  # Asignar None por defecto
                        )
                        copia_item.save()
                        copiault.append(copia_item)
                    else:
                        error_message = f"No se encontró el libro con la clave: {codigolibro}"

                return redirect('/copia_agregada_excel')  # Redirigir a la página de copias después de cargar las copias

            except Exception as e:
                error_message = 'Se produjo un error al intentar cargar las copias desde el archivo Excel. Por favor, inténtalo de nuevo.'
                return render(request, 'error.html', {'mensaje': error_message})

    return render(request, 'tu_template_excel.html')

@login_required #verifica que este iniciada sesion
def eliminar_copia(request, pk):
    copia_obj = get_object_or_404(Copia, clavecopia=pk)


    if request.method == 'POST':
        copia_obj.delete()
        return redirect('/busqueda')

    return redirect('/busqueda')


def nuevo_prestamo(request):
    if request.method == 'POST':
        clave_alumno = request.POST.get('clave_alumno')
        clave_copia = request.POST.get('clave_copia')

        if not (clave_alumno and clave_copia):
            error_message = 'Por favor, completa todos los campos obligatorios.'
            return render(request, 'error.html', {'mensaje': error_message})

        if clave_alumno.isdigit():
            # Eliminar los ceros a la izquierda en el caso de que existan
            clave_alumno = str(int(clave_alumno))
        if clave_copia.isdigit():
            # Eliminar los ceros a la izquierda en el caso de que existan
            clave_copia = str(int(clave_copia))

        try:
            alumno = Alumno.objects.get(clave=clave_alumno)
            copia = Copia.objects.get(clavecopia=clave_copia)

            if copia.disponible:
                prestamo = Prestamo.objects.create(
                    clave_alumno=alumno,
                    clave_copia=copia,
                    activo=True,
                    regreso=datetime.now() + timedelta(days=7),
                    fecha_regreso=None,
                    fecha_creacion = datetime.now()
                )

                # Modificar disponibilidad de copia y clave de alumno en la copia
                copia.disponible = False
                copia.clavealumno = alumno
                copia.save()

                # Actualizar estado de sacalibro en alumno
                alumno.sacalibro = True
                alumno.save()
                prestault.clear()
                prestault.append(prestamo)
                return redirect('/prestamo_nuevo')
            else:
                error_message = 'La copia no está disponible.'
                return render(request, 'error.html', {'mensaje': error_message})
        except (Alumno.DoesNotExist, Copia.DoesNotExist):
            error_message = 'La clave de alumno o de copia no son válidas.'
            return render(request, 'error.html', {'mensaje': error_message})

    error_message = 'El método de solicitud no es válido.'
    return render(request, 'error.html', {'mensaje': error_message})

def completar_prestamo(request, pk):
    prestamo = Prestamo.objects.get(pk=pk)
    prestamo.activo = False
    prestamo.fecha_regreso = timezone.now().date()
    prestamo.save()

    copia = prestamo.clave_copia
    copia.disponible = True
    copia.clavealumno = None
    copia.save()

    alumno = prestamo.clave_alumno
    alumno.sacalibro = False
    alumno.save()

    return redirect('/retorno')

def ampliar_prestamo(request, pk):
    prestamo = Prestamo.objects.get(pk=pk)
    prestamo.regreso += timezone.timedelta(days=7)
    prestamo.save()
    return redirect('/retorno')

@login_required #verifica que este iniciada sesion
def exportar_excel(request):
    try:
        # Crear un libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Copias"

        # Encabezados de la tabla
        ws.append(["Codigo de Copia", "Codigo de Libro","Titulo del Libro", "Autor", "Ilustrador", "Editorial", "Clasificacion Dewey"])

        # Obtener todas las copias
        copias = Copia.objects.all()

        # Datos de la tabla
        for copia in copias:
            ws.append([copia.clavecopia, copia.codigolibro.codigolibro ,copia.codigolibro.titulo, copia.codigolibro.autor, copia.codigolibro.ilustrador, copia.codigolibro.editorial, copia.codigolibro.dewy])

        # Guardar el libro de trabajo como un archivo Excel
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="copias.xlsx"'
        wb.save(response)
        
        # Retorna la respuesta
        return response
    
    except Exception as e:
        # Maneja la excepción imprimiendo el mensaje de error
        print(f"Error en la función exportar_excel: {e}")
        
        # Puedes redirigir a una página de error o simplemente retornar un HttpResponse con un mensaje de error
        return HttpResponse("Ocurrió un error al exportar los datos a Excel.")

@login_required #verifica que este iniciada sesion
def exportar_excel_alumnos(request):
    # Crear un libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Alumnos"

    # Encabezados de la tabla
    ws.append(["Clave", "Nombre", "Apellido", "Año"])

    # Obtener todos los alumnos
    alumnos = Alumno.objects.all()

    # Datos de la tabla
    for alumno in alumnos:
        ws.append([alumno.clave, alumno.nombre, alumno.apellido, alumno.grupo])

    # Guardar el libro de trabajo como un archivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="alumnos.xlsx"'
    wb.save(response)
    
    return response

@login_required #verifica que este iniciada sesion
def exportar_excel_libros(request):
    # Crear un libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Libros"

    # Encabezados de la tabla
    ws.append(["Código de Libro", "Título", "Autor", "Ilustrador", "Fecha de Publicación", "Editorial", "Número de Tomo", "Características Especiales", "Clasificación Dewey", "Público Dirigido"])

    # Obtener todos los libros
    libros = Libro.objects.all()

    # Datos de la tabla
    for libro in libros:
        ws.append([
            libro.codigolibro,
            libro.titulo,
            libro.autor,
            libro.ilustrador,
            libro.fechapublicacion,
            libro.editorial,
            libro.numerotomo,
            libro.caracteristicasespeciales,
            libro.dewy,
            libro.publicodirigido,
        ])

    # Guardar el libro de trabajo como un archivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="libros.xlsx"'
    wb.save(response)
    
    return response


@login_required #verifica que este iniciada sesion
def exportar_excel_prestamos(request):
    # Crear un libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Préstamos"

    # Encabezados de la tabla
    ws.append(["Clave de Alumno", "Nombre Alumno", "Clave de Libro", "Título libro", "Fecha límite", "Fecha regresado"])

    # Obtener todos los préstamos
    prestamos = Prestamo.objects.all()

    # Datos de la tabla
    for prestamo in prestamos:
        ws.append([
            prestamo.clave_alumno.clave,
            f"{prestamo.clave_alumno.nombre} {prestamo.clave_alumno.apellido}",
            prestamo.clave_copia.clavecopia,
            prestamo.clave_copia.codigolibro.titulo,
            prestamo.regreso,
            prestamo.fecha_regreso if prestamo.fecha_regreso else "",
        ])

    # Guardar el libro de trabajo como un archivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="prestamos_completos.xlsx"'
    wb.save(response)
    
    return response

@login_required #verifica que este iniciada sesion
def pagar_multa(request, pk):
    if request.method == 'POST':
        try:
            multa = Multa.objects.get(pk=pk)
            multa.pagado = True
            multa.save()
            return JsonResponse({'success': True})
        except Multa.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'La multa no existe'})
    else:
        return JsonResponse({'success': False, 'error': 'Método de solicitud incorrecto'})
    
@login_required #verifica que este iniciada sesion
def exportar_excel_multas(request):
    # Crear un libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Multas"

    # Encabezados de la tabla
    ws.append(["Clave de Alumno", "Nombre Alumno", "Año Alumno","Grupo", "Cantidad de multa", "Pagado"])

    # Obtener todas las multas
    multas = Multa.objects.all()

    # Datos de la tabla
    for multa in multas:
        ws.append([
            multa.alumno.clave,
            f"{multa.alumno.nombre} {multa.alumno.apellido}",
            "PF" if multa.alumno.grupo == 0 else multa.alumno.grupo  ,
            multa.alumno.clase,
            multa.monto,
            "Sí" if multa.pagado else "No",
        ])

    # Guardar el libro de trabajo como un archivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="multas.xlsx"'
    wb.save(response)
    
    return response


@login_required #verifica que este iniciada sesion
def exportar_excel_prestamos_grupo(request):
    # Crear un libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active
    
    query = request.POST.get('clave_alum')
    print(query)
    prestamos = None  # Inicializar la variable prestamos
    ws.title = "Préstamos de " + query
    # Encabezados de la tabla
    ws.append(["Clave de Alumno", "Nombre Alumno", "Clave de Libro", "Título libro", "Fecha límite", "Fecha regresado"])

    if query:
        if query[0]==0:
             ws.title = "Préstamos de PF" + query[1]
        else:
            ws.title = "Préstamos de " + query
        # Verificar si la consulta es un número puro
        prestamos = Prestamo.objects.annotate(grado_grupo=Concat('clave_alumno__grupo', 'clave_alumno__clase', output_field=CharField())).filter(grado_grupo=query, activo=True)


    if prestamos:  # Verificar si prestamos tiene un valor asignado
        # Datos de la tabla
        for prestamo in prestamos:
            ws.append([
                prestamo.clave_alumno.clave,
                f"{prestamo.clave_alumno.nombre} {prestamo.clave_alumno.apellido}",
                prestamo.clave_copia.clavecopia,
                prestamo.clave_copia.codigolibro.titulo,
                prestamo.regreso,
                prestamo.fecha_regreso if prestamo.fecha_regreso else "",
            ])

    # Guardar el libro de trabajo como un archivo Excel
    # Guardar el libro de trabajo como un archivo Excel
    nombre_archivo = f"{ws.title}.xlsx"  # Utiliza el título del libro de trabajo en el nombre del archivo
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)

    return response

@login_required #verifica que este iniciada sesion
def exportar_excel_prestamos_alumno(request):
    # Crear un libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active
    
    query = request.POST.get('clave_alum')
    print(query)
    prestamos = None  # Inicializar la variable prestamos
    ws.title = "Préstamos de " + query 
    # Encabezados de la tabla
    ws.append(["Clave de Alumno", "Nombre Alumno", "Clave de Libro", "Título libro", "Fecha límite", "Fecha regresado"])

    if query:
        # Verificar si la consulta es un número puro
        if query.isdigit():
            # Eliminar los ceros a la izquierda en el caso de que existan
            query = str(int(query))
        # Obtener todos los préstamos
        prestamos = Prestamo.objects.filter(clave_alumno__clave=query)

    if prestamos:  # Verificar si prestamos tiene un valor asignado
        # Datos de la tabla
        for prestamo in prestamos:
            ws.append([
                prestamo.clave_alumno.clave,
                f"{prestamo.clave_alumno.nombre} {prestamo.clave_alumno.apellido}",
                prestamo.clave_copia.clavecopia,
                prestamo.clave_copia.codigolibro.titulo,
                prestamo.regreso,
                prestamo.fecha_regreso if prestamo.fecha_regreso else "",
            ])

    # Guardar el libro de trabajo como un archivo Excel
    nombre_archivo = f"{ws.title}.xlsx"  # Utiliza el título del libro de trabajo en el nombre del archivo
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)

    
    return response


# Vista para agregar copia
@login_required #verifica que este iniciada sesion
def agregar_copia_todos(request):
    if request.method == 'POST':
        copia = request.POST.get('copia')
        if copia.isdigit():
            clave_copia = int(copia)
            # Verificar si ya existe una copia asociada a esa clave
            copia_existente = Copia.objects.filter(clavecopia=clave_copia).first()
            if copia_existente:
                # Si existe, agregarla a la lista global
                copia_front=auxiliar(copia_existente,1)
                copia_tras=auxiliar(copia_existente,2)
                copia_int=auxiliar(copia_existente,3)
                listacopiasfront.append(copia_front)
                listacopiastras.append(copia_tras)
                listacopiasint.append(copia_int)
    return redirect('/etiqueta')

@login_required #verifica que este iniciada sesion
def agregar_copia_int(request):
    if request.method == 'POST':
        copia = request.POST.get('copia')
        if copia.isdigit():
            clave_copia = int(copia)
            # Verificar si ya existe una copia asociada a esa clave
            copia_existente = Copia.objects.filter(clavecopia=clave_copia).first()
            if copia_existente:
                # Si existe, agregarla a la lista global
                copia_int=auxiliar(copia_existente,3)
                listacopiasint.append(copia_int)
    return redirect('/etiqueta')

@login_required #verifica que este iniciada sesion
def agregar_copia_front(request):
    if request.method == 'POST':
        copia = request.POST.get('copia')
        if copia.isdigit():
            clave_copia = int(copia)
            # Verificar si ya existe una copia asociada a esa clave
            copia_existente = Copia.objects.filter(clavecopia=clave_copia).first()
            if copia_existente:
                # Si existe, agregarla a la lista global
                copia_front=auxiliar(copia_existente,1)
                listacopiasfront.append(copia_front)
    return redirect('/etiqueta')

@login_required #verifica que este iniciada sesion
def agregar_copia_tras(request):
    if request.method == 'POST':
        copia = request.POST.get('copia')
        if copia.isdigit():
            clave_copia = int(copia)
            # Verificar si ya existe una copia asociada a esa clave
            copia_existente = Copia.objects.filter(clavecopia=clave_copia).first()
            if copia_existente:
                # Si existe, agregarla a la lista global
                copia_tras=auxiliar(copia_existente,2)
                listacopiastras.append(copia_tras)
    return redirect('/etiqueta')


# Vista para vaciar lista
@login_required #verifica que este iniciada sesion
def vaciar_lista(request):
    if request.method == 'POST':
        vaciocopias.clear()
        listacopiasfront.clear()
        listacopiastras.clear()
        listacopiasint.clear()
        return redirect('/etiqueta')  # Redirigir a donde desees después de vaciar la lista
    return redirect('/etiqueta')  # Renderizar tu template


@login_required #verifica que este iniciada sesion
def agregar_alu(request):
    if request.method == 'POST':
        copia = request.POST.get('alumno')
        if copia.isdigit():
            clave_copia = int(copia)
            # Verificar si ya existe una copia asociada a esa clave
            copia_existente = Alumno.objects.filter(clave=clave_copia).first()
            if copia_existente:
                # Si existe, agregarla a la lista global
                listacredenciales.append(copia_existente)
    return redirect('/credencial')

# Vista para vaciar lista
@login_required #verifica que este iniciada sesion
def vaciar_lista_alum(request):
    if request.method == 'POST':
        listacredenciales.clear()
        vacioalum.clear()
        return redirect('/credencial')
    return redirect('/credencial')


@login_required #verifica que este iniciada sesion
def quitar_registro_front(request, copia_id):
    for instance in listacopiasfront:
        if instance.copia.clavecopia == copia_id:
            listacopiasfront.remove(instance)
            break
    return redirect('/etiqueta')

@login_required #verifica que este iniciada sesion
def quitar_registro_back(request, copia_id):
    for instance in listacopiastras:
        if instance.copia.clavecopia == copia_id:
            listacopiastras.remove(instance)
            break
    return redirect('/etiqueta')

@login_required #verifica que este iniciada sesion
def quitar_registro_int(request, copia_id):
    for instance in listacopiasint:
        if instance.copia.clavecopia == copia_id:
            listacopiasint.remove(instance)
            break
    return redirect('/etiqueta')

@login_required #verifica que este iniciada sesion
def quitar_registro_alum(request, alumno_id):
    # Encuentra el objeto copia por su ID
    copia = Alumno.objects.get(clave=alumno_id)
    listacredenciales.remove(copia)
    return redirect('/credencial')

@login_required #verifica que este iniciada sesion
def agregar_alu_vacia(request):
    if request.method == 'POST':
        copia = request.POST.get('alumno')
        if copia.isdigit():
           clave_copia = int(copia)
           vacioalum.clear()
           for i in range(clave_copia):
               vacioalum.append(" ")
    return redirect('/credencial')

@login_required #verifica que este iniciada sesion
def agregar_copia_vacia(request):
    if request.method == 'POST':
        copia = request.POST.get('copia')
        if copia.isdigit():
            clave_copia = int(copia)
            vaciocopias.clear()
            for i in range(clave_copia):
                vacio= auxiliar("",0)
                vaciocopias.append(vacio)
    return redirect('/etiqueta')





def exportar_tabla(request):
    pass
    